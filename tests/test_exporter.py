"""
tests/test_exporter.py
Tests para src/exporter.py. No requieren credenciales de Schwab.
"""

import openpyxl
import pandas as pd
import pytest
from datetime import date

from src.exporter import export_to_excel, ITM_FILL_COLOR


EXPIRATION = date(2025, 6, 20)


def _minimal_calls(itm_flags=None):
    """DataFrame minimo de calls para tests."""
    itm_flags = itm_flags or [False, False]
    return pd.DataFrame({
        "strike":      [500.0, 505.0],
        "bid":         [1.0,   0.8],
        "ask":         [1.2,   1.0],
        "inTheMoney":  itm_flags,
    })


def _minimal_puts(itm_flags=None):
    """DataFrame minimo de puts para tests."""
    itm_flags = itm_flags or [False, False]
    return pd.DataFrame({
        "strike":     [500.0, 505.0],
        "bid":        [0.5,   0.6],
        "ask":        [0.7,   0.8],
        "inTheMoney": itm_flags,
    })


def test_export_creates_file(tmp_path, monkeypatch):
    """export_to_excel debe generar un archivo .xlsx en OUTPUT_DIR."""
    monkeypatch.setattr("config.OUTPUT_DIR", str(tmp_path))

    filepath = export_to_excel(_minimal_calls(), _minimal_puts(), "SPY", EXPIRATION)

    assert filepath.exists()
    assert filepath.suffix == ".xlsx"
    assert "SPY" in filepath.name
    assert "20250620" in filepath.name


def test_export_sheet_names(tmp_path, monkeypatch):
    """El Excel generado debe contener exactamente las sheets CALLS, PUTS e INFO."""
    monkeypatch.setattr("config.OUTPUT_DIR", str(tmp_path))

    filepath = export_to_excel(_minimal_calls(), _minimal_puts(), "SPY", EXPIRATION)

    wb = openpyxl.load_workbook(filepath)
    assert set(wb.sheetnames) == {"CALLS", "PUTS", "INFO"}


def test_itm_detection(tmp_path, monkeypatch):
    """Las filas con inTheMoney=True deben recibir el color ITM en la sheet CALLS."""
    monkeypatch.setattr("config.OUTPUT_DIR", str(tmp_path))

    # Primera fila es ITM, segunda no
    calls = _minimal_calls(itm_flags=[True, False])
    filepath = export_to_excel(calls, _minimal_puts(), "SPY", EXPIRATION)

    wb = openpyxl.load_workbook(filepath)
    ws = wb["CALLS"]

    # Encontrar indice de columna "inTheMoney"
    itm_col = None
    for cell in ws[1]:
        if cell.value == "inTheMoney":
            itm_col = cell.column
            break
    assert itm_col is not None, "Columna 'inTheMoney' no encontrada en el header"

    # Fila 2 (primera fila de datos) debe tener el color ITM
    itm_row_fill = ws.cell(row=2, column=1).fill.fgColor.rgb
    assert itm_row_fill.upper().endswith(ITM_FILL_COLOR.upper()), (
        f"Se esperaba color ITM {ITM_FILL_COLOR}, se obtuvo {itm_row_fill}"
    )

    # Fila 3 (segunda fila de datos) NO debe tener el color ITM
    non_itm_row_fill = ws.cell(row=3, column=1).fill.fgColor.rgb
    assert not non_itm_row_fill.upper().endswith(ITM_FILL_COLOR.upper()), (
        "Una fila no-ITM no debe tener el color ITM"
    )


def _analysis_fixture():
    return {
        "max_pain": {"strike": 500.0, "pain_by_strike": {500.0: 100.0, 505.0: 200.0}},
        "pc_ratio": {"volume_ratio": 1.25, "oi_ratio": 0.8750},
    }


def test_analysis_sheet_created_when_analysis_provided(tmp_path, monkeypatch):
    monkeypatch.setattr("config.OUTPUT_DIR", str(tmp_path))
    filepath = export_to_excel(
        _minimal_calls(), _minimal_puts(), "SPY", EXPIRATION,
        analysis=_analysis_fixture(),
    )
    wb = openpyxl.load_workbook(filepath)
    assert "ANALYSIS" in wb.sheetnames


def test_analysis_sheet_absent_without_analysis(tmp_path, monkeypatch):
    monkeypatch.setattr("config.OUTPUT_DIR", str(tmp_path))
    filepath = export_to_excel(_minimal_calls(), _minimal_puts(), "SPY", EXPIRATION)
    wb = openpyxl.load_workbook(filepath)
    assert "ANALYSIS" not in wb.sheetnames


def test_analysis_sheet_contains_max_pain(tmp_path, monkeypatch):
    monkeypatch.setattr("config.OUTPUT_DIR", str(tmp_path))
    filepath = export_to_excel(
        _minimal_calls(), _minimal_puts(), "SPY", EXPIRATION,
        analysis=_analysis_fixture(),
    )
    wb = openpyxl.load_workbook(filepath)
    ws = wb["ANALYSIS"]
    all_values = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert "$500.00" in all_values


def test_analysis_sheet_contains_pc_ratio(tmp_path, monkeypatch):
    monkeypatch.setattr("config.OUTPUT_DIR", str(tmp_path))
    filepath = export_to_excel(
        _minimal_calls(), _minimal_puts(), "SPY", EXPIRATION,
        analysis=_analysis_fixture(),
    )
    wb = openpyxl.load_workbook(filepath)
    ws = wb["ANALYSIS"]
    all_values = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert "1.2500" in all_values


def test_export_multiple_creates_file(tmp_path, monkeypatch):
    from src.exporter import export_multiple_to_excel
    monkeypatch.setattr("config.OUTPUT_DIR", str(tmp_path))
    exp1 = date(2025, 6, 20)
    exp2 = date(2025, 7, 18)
    parsed = {
        exp1: (_minimal_calls(), _minimal_puts()),
        exp2: (_minimal_calls(), _minimal_puts()),
    }
    skew_df = pd.DataFrame({"strike": [500.0], "2025-06-20": [0.20], "2025-07-18": [0.25]})
    filepath = export_multiple_to_excel(parsed, skew_df, "SPY")
    assert filepath.exists()
    assert filepath.suffix == ".xlsx"


def test_export_multiple_sheet_names(tmp_path, monkeypatch):
    from src.exporter import export_multiple_to_excel
    monkeypatch.setattr("config.OUTPUT_DIR", str(tmp_path))
    exp1 = date(2025, 6, 20)
    parsed = {exp1: (_minimal_calls(), _minimal_puts())}
    skew_df = pd.DataFrame({"strike": [500.0], "2025-06-20": [0.20]})
    filepath = export_multiple_to_excel(parsed, skew_df, "SPY")
    wb = openpyxl.load_workbook(filepath)
    assert "CALLS_20250620" in wb.sheetnames
    assert "PUTS_20250620" in wb.sheetnames
    assert "IV_SKEW" in wb.sheetnames
    assert "INFO" in wb.sheetnames


def test_export_multiple_filename_contains_all_dates(tmp_path, monkeypatch):
    from src.exporter import export_multiple_to_excel
    monkeypatch.setattr("config.OUTPUT_DIR", str(tmp_path))
    exp1 = date(2025, 6, 20)
    exp2 = date(2025, 7, 18)
    parsed = {
        exp1: (_minimal_calls(), _minimal_puts()),
        exp2: (_minimal_calls(), _minimal_puts()),
    }
    skew_df = pd.DataFrame({"strike": [500.0]})
    filepath = export_multiple_to_excel(parsed, skew_df, "SPY")
    assert "20250620" in filepath.name
    assert "20250718" in filepath.name
