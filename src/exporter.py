"""
src/exporter.py
Exporta los DataFrames de calls y puts a un archivo .xlsx
con formato visual claro para el analista financiero.

Genera:
  output/{SYMBOL}_{YYYYMMDD}_options.xlsx
    |- Sheet "CALLS"
    |- Sheet "PUTS"
    +- Sheet "INFO"  <- metadatos de la descarga
"""

import pathlib
from datetime import date, datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config


# -- Colores -------------------------------------------------
CALL_HEADER_COLOR = "1F4E79"   # Azul oscuro
PUT_HEADER_COLOR  = "7B2D00"   # Rojo oscuro
INFO_HEADER_COLOR = "2E4057"   # Gris azulado
HEADER_FONT_COLOR = "FFFFFF"   # Blanco
ITM_FILL_COLOR    = "E8F5E9"   # Verde claro (In The Money)
ALT_ROW_COLOR     = "F5F5F5"   # Gris muy claro (filas alternas)


def _style_sheet(ws: Worksheet, header_color: str) -> None:
    """Aplica formato al worksheet."""
    header_fill = PatternFill("solid", fgColor=header_color)
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=10)
    itm_fill    = PatternFill("solid", fgColor=ITM_FILL_COLOR)
    alt_fill    = PatternFill("solid", fgColor=ALT_ROW_COLOR)
    center      = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD")
    )

    # Header
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center

    # Encontrar el indice de la columna "inTheMoney" una sola vez
    itm_col_idx = None
    for cell in ws[1]:
        if cell.value == "inTheMoney":
            itm_col_idx = cell.column
            break

    # Filas de datos
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        is_itm = False
        if itm_col_idx is not None:
            val = ws.cell(row=row_idx, column=itm_col_idx).value
            is_itm = val is True or str(val).lower() == "true"

        for cell in row:
            cell.alignment = center
            cell.border    = thin_border
            if is_itm:
                cell.fill = itm_fill
            elif row_idx % 2 == 0:
                cell.fill = alt_fill

    # Ancho de columnas automatico
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    # Freeze pane en la primera fila
    ws.freeze_panes = "A2"


def _write_sheet(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    sheet_name: str,
    header_color: str,
) -> None:
    """Escribe un DataFrame como sheet y aplica estilos."""
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    _style_sheet(writer.sheets[sheet_name], header_color)


def _write_analysis_sheet(
    writer: pd.ExcelWriter,
    analysis: dict,
    calls_df: pd.DataFrame,
    puts_df: pd.DataFrame,
) -> None:
    max_pain = analysis.get("max_pain", {})
    pc_ratio = analysis.get("pc_ratio", {})

    mp_strike = max_pain.get("strike")
    vol_ratio = pc_ratio.get("volume_ratio")
    oi_ratio  = pc_ratio.get("oi_ratio")

    data = {
        "Metrica": [
            "Max Pain Strike",
            "P/C Ratio (Volumen)",
            "P/C Ratio (Open Interest)",
            "Total Calls",
            "Total Puts",
        ],
        "Valor": [
            f"${mp_strike:.2f}" if isinstance(mp_strike, (int, float)) else "N/A",
            f"{vol_ratio:.4f}"  if isinstance(vol_ratio, float) and vol_ratio != float("inf") else str(vol_ratio),
            f"{oi_ratio:.4f}"   if isinstance(oi_ratio,  float) and oi_ratio  != float("inf") else str(oi_ratio),
            str(len(calls_df)),
            str(len(puts_df)),
        ],
        "Interpretacion": [
            "Strike donde la mayoria de opciones expira worthless",
            "> 1.0 bearish  /  < 1.0 bullish",
            "> 1.0 bearish  /  < 1.0 bullish",
            "",
            "",
        ],
    }
    _write_sheet(writer, pd.DataFrame(data), "ANALYSIS", INFO_HEADER_COLOR)


def export_to_excel(
    calls_df: pd.DataFrame,
    puts_df: pd.DataFrame,
    symbol: str,
    expiration: date,
    analysis: dict | None = None,
) -> pathlib.Path:
    """
    Genera el archivo Excel con las sheets CALLS, PUTS e INFO.

    Args:
        calls_df:   DataFrame de calls devuelto por parse_option_chain().
        puts_df:    DataFrame de puts devuelto por parse_option_chain().
        symbol:     Ticker del subyacente (ej: "SPY"). Se convierte a uppercase.
        expiration: Fecha de vencimiento usada para nombrar el archivo.

    Returns:
        Path absoluto del archivo .xlsx generado en OUTPUT_DIR.
    """
    output_dir = pathlib.Path(config.OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"{symbol.upper()}_{expiration.strftime('%Y%m%d')}_options.xlsx"
    filepath = output_dir / filename

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        _write_sheet(writer, calls_df, "CALLS", CALL_HEADER_COLOR)
        _write_sheet(writer, puts_df,  "PUTS",  PUT_HEADER_COLOR)

        if analysis is not None:
            _write_analysis_sheet(writer, analysis, calls_df, puts_df)

        # -- INFO -------------------------------------------
        info_data = {
            "Campo": [
                "Subyacente",
                "Vencimiento",
                "Descargado el",
                "Calls encontradas",
                "Puts encontradas",
                "Fuente",
            ],
            "Valor": [
                symbol.upper(),
                expiration.strftime("%Y-%m-%d"),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                len(calls_df),
                len(puts_df),
                "Schwab Market Data API",
            ],
        }
        _write_sheet(writer, pd.DataFrame(info_data), "INFO", INFO_HEADER_COLOR)

    print(f"[OK] Excel generado: {filepath}")
    return filepath


def export_multiple_to_excel(
    parsed: dict,
    skew_df: pd.DataFrame,
    symbol: str,
) -> pathlib.Path:
    """
    Genera el archivo Excel para multiples vencimientos.

    Args:
        parsed:  {expiration: (calls_df, puts_df)} de parse_option_chain().
        skew_df: DataFrame de IV Skew de calculate_iv_skew().
        symbol:  Ticker del subyacente.

    Returns:
        Path absoluto del archivo .xlsx generado en OUTPUT_DIR.
    """
    output_dir = pathlib.Path(config.OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)

    sorted_exps = sorted(parsed.keys())
    dates_str   = "_".join(exp.strftime("%Y%m%d") for exp in sorted_exps)
    filename    = f"{symbol.upper()}_{dates_str}_options.xlsx"
    filepath    = output_dir / filename

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for exp in sorted_exps:
            calls_df, puts_df = parsed[exp]
            tag = exp.strftime("%Y%m%d")
            _write_sheet(writer, calls_df, f"CALLS_{tag}", CALL_HEADER_COLOR)
            _write_sheet(writer, puts_df,  f"PUTS_{tag}",  PUT_HEADER_COLOR)

        if not skew_df.empty:
            _write_sheet(writer, skew_df, "IV_SKEW", INFO_HEADER_COLOR)

        total_calls = sum(len(parsed[e][0]) for e in sorted_exps)
        total_puts  = sum(len(parsed[e][1]) for e in sorted_exps)
        exps_str    = ", ".join(exp.strftime("%Y-%m-%d") for exp in sorted_exps)

        info_data = {
            "Campo": [
                "Subyacente", "Vencimientos", "Descargado el",
                "Total Calls", "Total Puts", "Fuente",
            ],
            "Valor": [
                symbol.upper(),
                exps_str,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                total_calls,
                total_puts,
                "Schwab Market Data API",
            ],
        }
        _write_sheet(writer, pd.DataFrame(info_data), "INFO", INFO_HEADER_COLOR)

    print(f"[OK] Excel generado: {filepath}")
    return filepath
